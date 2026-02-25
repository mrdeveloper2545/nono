from django.shortcuts import render, redirect, get_list_or_404, get_object_or_404
from .models import Unit,Material,Category,Vendor,PurchaseItem,PurchaseOrder,Product,Order,OrderItem
from dashboard.models import Settings,Expenses,Service
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from datetime import datetime, timedelta, time
from django.contrib import messages
from django.db.models import Sum,F, ExpressionWrapper, DecimalField
from django.http import HttpResponse,HttpResponseForbidden
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
import json
from django.utils import timezone
from django.db.models.functions import TruncDate
from django.utils.dateparse import parse_date
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
from django.conf import settings
import os
from django.db import transaction
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo






# Create your views here.

# UNIT MANAGEMENT VIEW
class UnitManagement(View,PermissionRequiredMixin,LoginRequiredMixin):
    template_name='unit/management.html'
    permission_required='inventory.view_unit'

    def get(self,request,*args,**kwargs):
        settings = Settings.objects.first()
        messages_data = [{"message":message.message, 'tags':message.tags} for message in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)
        units=Unit.objects.all()
        now=datetime.now()
        context={
            'time':now,
            'year':now.year,
            'month':now.strftime('%B'),
            'day':now.strftime('%A'),
            'units':units,
            'settings':settings,
            'messages_json':messages_json,
        }
        return render(request, self.template_name,context)

    def post(self,request,*args,**kwargs):
        names = request.POST.getlist('name[]')

        if not names:
            messages.info(request, 'name are not provided')
            return redirect('unit-management')

        if Unit.objects.filter(name__in=names).exists():
            messages.info(request, 'One or more units with the given names already exists')
            return redirect('unit-management')

        units = []
        for name in names:
            unit = Unit(name=name,user=request.user)
            units.append(unit)

        Unit.objects.bulk_create(units)
        messages.success(request, 'units added successfully')
        return redirect('unit-management')

class UpdateDeleteUnit(View,PermissionRequiredMixin,LoginRequiredMixin):
    def get(self,request,id,*args,**kwargs):
        Unit.objects.get(id=id).delete()
        messages.success(request, 'unit deleted successfully')
        return redirect('unit-management')

    def post(self,request,id,*args,**kwargs):
        unit = Unit.objects.get(id=id)
        names=request.POST.getlist('name[]')

        for i in range(len(names)):
            name = names[i]


            existing_unit = Unit.objects.filter(name__in=names).exclude(id=unit.id).first()
            if existing_unit:
                messages.error(request, 'unit with that name already exist')
                return redirect('unit-management')


        if i == 0:
            unit.name = name
            unit.save()
        else:
            new_unit = Unit(name=name)
            new_unit.save()
            messages.success(request, 'unit updated successfully')
        return redirect('unit-management')

class BulkUpdateUnit(View, LoginRequiredMixin, PermissionRequiredMixin):
    permission_required = ''

    def post(self, request, *args, **kwargs):
        units = get_list_or_404(Unit)


        for unit in units:
            updated_name = request.POST.get(f'name_{unit.id}')

            # check existing unit
            existing_unit = Unit.objects.filter(name=updated_name).exclude(id=unit.id).first()
            if existing_unit:
                messages.error(request, f"unit with ID {unit.id} already have the same name")

            # Update after checking complete
            unit.name = updated_name
        Unit.objects.bulk_update(units, ['name'])
        messages.success(request, 'units updated successfully')
        return redirect('unit-management')


# CATEGORY MANAGEMENT VIEW
class CategoryManagement(PermissionRequiredMixin, LoginRequiredMixin, View):
    template_name = 'category/management.html'
    permission_required = 'inventory.view_category'

    def get(self, request, *args, **kwargs):
        settings = Settings.objects.first()
        categories = Category.objects.all()
        now = datetime.now()

        # Serialize messages for the current request
        messages_data = [{'message': message.message, 'tags': message.tags} for message in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)

        context = {
            'categories': categories,
            'time': now,
            'year': now.year,
            'month': now.strftime('%B'),
            'day': now.strftime('%A'),
            'messages_json': messages_json,
            'settings':settings
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        names = request.POST.getlist('name[]')
        if not names:
            messages.error(request, 'Names are not provided')
            return redirect('category-management')

        if Category.objects.filter(name__in=names).exists():
            messages.error(request, 'One or category with the given names already exist')
            return redirect ('category-management')

        categories = []
        for name in names:
            category = Category(name=name,user=request.user)
            categories.append(category)

        Category.objects.bulk_create(categories)
        messages.success(request, 'Category created successfully')
        return redirect('category-management')

class UpdateDeleteCategory(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = ""

    def post(self, request, id, *args, **kwargs):
        category = get_object_or_404(Category, id=id)

        names = request.POST.getlist('name[]')
        for i in range(len(names)):

            existing_category = Category.objects.filter(name__in=names).exclude(id=category.id).first()
            if existing_category:
                messages.error(request, 'Category with name already exist')
                return redirect('category-management')

            if i == 0:
                category.name = names[i]
                category.save()
            else:
                new_category = Category(name=names[i])
                new_category.save()
                messages.success(request, 'category updated successfully')
        return redirect('category-management')

    def get(self, request, id, *args, **kwargs):
        Category.objects.get(id=id).delete()
        messages.success(request, 'category deleted successfully')
        return redirect('category-management')

class BulkUpdateCategory(View, LoginRequiredMixin, PermissionRequiredMixin):
    permission_required = 'inventory.change_category'

    def post(self, request, *args, **kwargs):
        categories = get_list_or_404(Category)

        for category in categories:
            updated_category = request.POST.get(f'name_{category.id}')
            existing_category = Category.objects.filter(name=updated_category).exclude(id=category.id).first()
            if existing_category:
                messages.error(request, f'category with id {category.id} already exist')
            category.name = updated_category
        Category.objects.bulk_update(categories, ['name'])
        messages.success(request, 'categories updated successfully')
        return redirect('category-management')




# MATERIAL MANAGEMENT VIEW
class MaterialManagement(LoginRequiredMixin, PermissionRequiredMixin, View):

    template_name = 'material/management.html'
    permission_required = 'inventory.view_material'

    def get(self, request, *args, **kwargs):
        settings = Settings.objects.first()
        materials = Material.objects.all()
        units = Unit.objects.all()
        categories = Category.objects.all()
        now = timezone.datetime.now()
        messages_data = [{'messages':message.message,'tags':message.tags} for message in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)
        context={
            'materials':materials,
            'time':now,
            'messages_json':messages_json,
            'month':now.strftime('%B'),
            'day':now.strftime('%A'),
            'settings':settings,
            'units':units,
            'categories':categories,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        names = request.POST.getlist('name[]')
        categories = request.POST.getlist('category[]')
        units = request.POST.getlist('unit[]')

        # Validate that all lists are the same length
        if not (len(names) == len(categories) == len(units)):
            messages.error(request, 'All fields must have the same number of entries.')
            return redirect('material-management')

        # Check that names are provided
        if not names:
            messages.info(request, 'No materials provided.')
            return redirect('material-management')

        # Check for duplicate names
        existing_names = set(Material.objects.filter(name__in=names).values_list('name', flat=True))
        if existing_names:
            messages.info(request, f"The following materials already exist: {', '.join(existing_names)}")
            return redirect('material-management')

        materials = []
        for i in range(len(names)):
            category = get_object_or_404(Category, id=categories[i])
            unit = get_object_or_404(Unit, id=units[i])
            user=request.user

            material = Material(
                name=names[i],
                category=category,
                unit=unit,
                user=user
            )
            materials.append(material)

        # Bulk create all materials
        Material.objects.bulk_create(materials)
        messages.success(request, 'Materials added successfully.')
        return redirect('material-management')


class UpdateDeleteMaterial(View, LoginRequiredMixin, PermissionRequiredMixin):
    permission_required = ''

    def post(self, request, id, *args, **kwargs):
        material = Material.objects.get(id=id)

        names = request.POST.getlist('name[]')

        for i in range(len(names)):


            existing_material = Material.objects.filter(name__in=names).exclude(id=material.id).first()
            if existing_material:
                continue


            if i == 0:
                material.name = names[i]
                material.save()

            else:

                new_material = Material(
                    name = names[i]
                )
                new_material.save()
                messages.success(request, 'material updated successfully')
        return redirect('material-management')



    def get(self, request, id, *args, **kwargs):
        Material.objects.get(id=id).delete()
        messages.success(request, 'material deleted successfully')
        return redirect('material-management')

class BulkUpdateMaterial(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = ''
    def post(self, request, *args, **kwargs):
        materials = get_list_or_404(Material)

        for material in materials:
            update_material = request.POST.get(f'name_{material.id}')
            material.name = update_material
        Material.objects.bulk_update(materials, ['name'])
        messages.success(request, 'Materials updated successfully')
        return redirect('material-management')



# VENDOR MANAGEMENT VIEW
class VendorManagement(View, LoginRequiredMixin, PermissionRequiredMixin):
    template_name = 'vendor/vendor.html'
    permission_required = 'inventory.view_vendor'

    def get(self, request, *args, **kwargs):
        now = timezone.datetime.now()
        messages_data = [{'messages':message.message,'tags':message.tags} for message in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)
        vendors = Vendor.objects.all()
        settings = Settings.objects.first()
        materials = Material.objects.all()

        context = {
            'time':now,
            'year':now.year,
            'month':now.strftime('%B'),
            'day':now.strftime('%A'),
            'messages_json':messages_json,
            'vendors':vendors,
            'settings':settings,
            'materials':materials
        }
        return render(request, self.template_name,context)


    def post(self, request, *args, **kwargs):
        user = request.POST.get('user')
        names = request.POST.getlist('name[]')
        phones = request.POST.getlist('phone_number[]')

        user = request.user
        vendors = []
        for name,phone_number in zip(names,phones):
            vendors.append(Vendor(name=name,user=user,phone_number=phone_number))
        Vendor.objects.bulk_create(vendors)
        print('created')
        return redirect('vendor-management')

class BulkUpdateVendors(View, LoginRequiredMixin, PermissionRequiredMixin):
    permission_required = ''

    def post(self, request, *args, **kwargs):
        vendors = get_list_or_404(Vendor)

        for vendor in vendors:
            vendor.name = request.POST.get(f'name_{vendor.id}')
            vendor.phone_number = request.POST.get(f'phone_number_{vendor.id}')


        Vendor.objects.bulk_update(vendors, ['name','phone_number'])
        return redirect('vendor-management')


class UpdateDeleteVendor(View, LoginRequiredMixin, PermissionRequiredMixin):
    permission_required = ''

    def get(self, request, id, *args, **kwargs):
        Vendor.objects.get(id=id).delete()
        messages.success(request, 'vendor deleted successfully')
        return redirect('vendor-management')


    def post(self, request, id, *args, **kwargs):
        name = request.POST.get('name')
        phone_number = request.POST.get('phone_number')
        Vendor.objects.filter(id=id).update(name=name,phone_number=phone_number)
        messages.success(request, 'vendor updated successfully')
        return redirect('vendor-management')


# PURCHASEORDERMANAGEMENT

class PurchaseManagementView(PermissionRequiredMixin, LoginRequiredMixin, View):
    template_name = 'purchase/management.html'
    permission_required = 'inventory.view_purchaseorder'

    def get(self, request, *args, **kwargs):
        now = timezone.now()
        messages_data = [{'messages': message.message, 'tags': message.tags} for message in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)

        context = {
            'time': now,
            'year': now.year,
            'month': now.strftime('%B'),
            'day': now.strftime('%A'),
            'messages_json': messages_json,
            'settings': Settings.objects.first(),
            'vendors': Vendor.objects.all(),
            'materials': Material.objects.all(),
            'po': PurchaseOrder.objects.all().order_by('-id'),
        }

        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        vendor_ids = request.POST.getlist('vendor[]')
        material_ids = request.POST.getlist('material[]')
        quantities = request.POST.getlist('quantity[]')
        prices = request.POST.getlist('price[]')
        order_date_str = request.POST.get('order_date')

        if not order_date_str:
            messages.error(request, "Order date is required.")
            return redirect('purchase-management')

        if not (len(vendor_ids) == len(material_ids) == len(quantities) == len(prices)):
            messages.error(request, "Mismatch in purchase items data.")
            return redirect('purchase-management')

        try:
            order_date = datetime.strptime(order_date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Invalid order date format.")
            return redirect('purchase-management')

        # Bulk fetch for efficiency
        materials = Material.objects.in_bulk([int(m) for m in material_ids])
        vendors = Vendor.objects.in_bulk([int(v) for v in vendor_ids])

        try:
            with transaction.atomic():
                po = PurchaseOrder(order_date=order_date, status='pending', user=request.user)
                po.save()

                created_items = 0

                for i in range(len(material_ids)):
                    try:
                        material = materials.get(int(material_ids[i]))
                        vendor = vendors.get(int(vendor_ids[i]))
                        quantity = int(quantities[i])
                        price = Decimal(prices[i])

                        if not material or not vendor:
                            raise ValueError("Invalid material or vendor reference.")
                        if quantity <= 0 or price < 0:
                            raise ValueError("Quantity must be positive and price non-negative.")

                        PurchaseItem.objects.create(
                            purchase_order=po,
                            material=material,
                            vendor=vendor,
                            quantity=quantity,
                            price=price,
                            user=request.user
                        )
                        created_items += 1

                    except (ValueError, InvalidOperation) as e:
                        messages.warning(request, f"Skipping item #{i + 1}: {e}")

                if created_items == 0:
                    po.delete()
                    messages.error(request, "No valid purchase items created. Purchase order cancelled.")
                    return redirect('purchase-management')

                messages.success(request, f"Purchase Order {po.order_number} created with {created_items} items.")
                return redirect('purchase-management')

        except Exception as e:
            messages.error(request, f"Failed to create Purchase Order: {e}")
            return redirect('purchase-management')


from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from .models import PurchaseOrder, PurchaseItem, Vendor, Material
from datetime import datetime

class UpdateDeletePurchaseOrderView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'order.change_purchaseorder'

    def post(self, request, *args, **kwargs):
        po_id = kwargs.get('pk')
        po = get_object_or_404(PurchaseOrder, id=po_id)

        # DELETE ORDER
        if request.POST.get('delete_order') == '1':
            if not request.user.has_perm('order.delete_purchaseorder'):
                messages.error(request, "You do not have permission to delete this order.")
                return redirect('purchase-management')
            try:
                with transaction.atomic():
                    # If approved, reduce stock
                    if po.status == 'approved':
                        for item in po.items.all():
                            item.material.quantity -= item.quantity
                            item.material.save()
                    po.delete()
                messages.success(request, f"Purchase Order {po.order_number} deleted successfully.")
            except Exception as e:
                messages.error(request, f"Failed to delete Purchase Order: {e}")
            return redirect('purchase-management')

        # UPDATE ORDER
        order_date_str = request.POST.get('order_date')
        vendor_ids = request.POST.getlist('vendor[]')
        material_ids = request.POST.getlist('material[]')
        quantities = request.POST.getlist('quantity[]')
        prices = request.POST.getlist('price[]')
        item_ids = request.POST.getlist('item_id[]')

        # Validate order date
        if order_date_str:
            try:
                po.order_date = datetime.strptime(order_date_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Invalid order date format.")
                return redirect('purchase-management')

        try:
            with transaction.atomic():
                po.save()
                existing_item_ids = set(po.items.values_list('id', flat=True))
                submitted_item_ids = set()

                for i in range(len(material_ids)):
                    # Skip incomplete rows
                    if not vendor_ids[i] or not material_ids[i] or not quantities[i] or not prices[i]:
                        continue

                    item_id = item_ids[i].strip()
                    vendor = get_object_or_404(Vendor, id=vendor_ids[i])
                    material = get_object_or_404(Material, id=material_ids[i])

                    # Convert quantity and price safely
                    try:
                        quantity = int(quantities[i])
                        price = Decimal(prices[i]).quantize(Decimal('0.01'))
                    except (ValueError, InvalidOperation):
                        messages.error(request, f"Invalid quantity or price at row {i+1}.")
                        return redirect('purchase-management')

                    # EXISTING ITEM
                    if item_id:
                        submitted_item_ids.add(int(item_id))
                        item = get_object_or_404(PurchaseItem, id=item_id, purchase_order=po)

                        # If PO approved, adjust stock for quantity change
                        if po.status == 'approved':
                            diff_qty = quantity - item.quantity
                            item.material.quantity += diff_qty
                            item.material.save()

                        # Update item fields
                        item.vendor = vendor
                        item.material = material
                        item.quantity = quantity
                        item.price = price
                        item.save()

                    # NEW ITEM
                    else:
                        item = PurchaseItem.objects.create(
                            purchase_order=po,
                            vendor=vendor,
                            material=material,
                            quantity=quantity,
                            price=price
                        )
                        submitted_item_ids.add(item.id)

                        # If PO approved, increment stock
                        if po.status == 'approved':
                            item.material.quantity += quantity
                            item.material.save()

                # DELETE REMOVED ITEMS
                to_delete = existing_item_ids - submitted_item_ids
                for item in PurchaseItem.objects.filter(id__in=to_delete):
                    if po.status == 'approved':
                        item.material.quantity -= item.quantity
                        item.material.save()
                    item.delete()

                # Update total cost
                po.update_total_cost()

                messages.success(request, f"Purchase Order {po.order_number} updated successfully.")
                return redirect('purchase-management')

        except Exception as e:
            messages.error(request, f"Failed to update Purchase Order: {e}")
            return redirect('purchase-management')


class PurchaseOrderDetailView(View, PermissionRequiredMixin, LoginRequiredMixin):
    template_name = 'purchase/detail.html'
    def get(self, request, id, *args, **kwargs):
        po = PurchaseOrder.objects.get(id=id)
        now = timezone.datetime.now()
        messages_data = [{'messages':message.message,'tags':message.tags} for message in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)
        settings = Settings.objects.first()
        context = {
            'time':now,
            'year':now.year,
            'month':now.strftime('%B'),
            'day':now.strftime('%A'),
            'messages_json':messages_json,
            'settings':settings,
            'po':po,
        }
        return render(request, self.template_name, context)


class ReceivePurchaseOrder(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'purchase.change_purchase_order'

    def get(self, request, id, *args, **kwargs):
        po = get_object_or_404(PurchaseOrder, id=id)

        if po.status == 'approved':
            messages.info(request, f'Purchase order {po.order_number} has already been received.')
        elif po.status == 'cancelled':
            messages.warning(request, f'Purchase order {po.order_number} is cancelled and cannot be received.')
        else:
            try:
                with transaction.atomic():
                    po.status = 'approved'
                    po.received_date = timezone.now()
                    po.save()  # This will also update material quantities if implemented in PO.save()
                    messages.success(request, f'Purchase order {po.order_number} received successfully.')
            except Exception as e:
                messages.error(request, f'Failed to receive purchase order: {e}')

        return redirect('purchase-management')
    

class CancellPurchaseOrder(View, PermissionRequiredMixin, LoginRequiredMixin):
    permission_required = 'change_purchase_order'

    def get(self, request, id, *args, **kwargs):
        po = PurchaseOrder.objects.get(id=id)
        if po.status != 'cancelled':
            po.status = 'cancelled'
            po.cancelled_date = timezone.now()
            po.save()
            messages.success(request, f'purchase order with number {po.order_number} cancelled successfully')
        return redirect('purchase-management')
    
class SinglePurchaseReport(LoginRequiredMixin, PermissionRequiredMixin, View):
    template_name = 'purchase/detail_report.html'
    permission_required = 'purchase.view_purchaseorder'

    def get(self, request, id, *args, **kwargs):
        # 1. Fetch the PurchaseOrder or return 404
        purchase = get_object_or_404(PurchaseOrder, id=id)

        # 2. Render HTML template with context
        context = {
            'purchase': purchase,
            'request': request,  # required for absolute static/media URLs
        }
        html = render_to_string(self.template_name, context)

        # 3. Create PDF from HTML
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="purchase-Invoice.pdf"'

        # Create a BytesIO buffer
        result = BytesIO()

        # Convert HTML to PDF
        pdf = pisa.pisaDocument(BytesIO(html.encode('utf-8')), dest=result,
            encoding='UTF-8', link_callback=self.link_callback)

        if pdf.err:
            return HttpResponse(f"Error generating PDF: <pre>{pdf.err}</pre>")

        # 4. Write result to response
        response.write(result.getvalue())
        return response

    def link_callback(self, uri, rel):
        """
        Convert HTML URIs to absolute system paths so xhtml2pdf can access those resources (e.g. images, CSS).
        """
        if uri.startswith(settings.MEDIA_URL):
            path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
        elif uri.startswith(settings.STATIC_URL):
            path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
        else:
            return uri  # return unchanged

        if not os.path.isfile(path):
            raise Exception(f'Media path does not exist: {path}')
        return path

class PendingPurchaseOrder(LoginRequiredMixin, PermissionRequiredMixin, View):
    template_name = 'purchase/pending.html'
    permission_required = 'inventory.view_purchase_order'

    def get(self, request, *args, **kwargs):
        now = timezone.datetime.now()
        messages_data = [{'messages':message.message,'tags':message.tags} for message in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)
        settings = Settings.objects.first()
        po = PurchaseOrder.objects.filter(status='pending')
        context = {
            'time':now,
            'year':now.year,
            'month':now.strftime('%B'),
            'day':now.strftime('%A'),
            'messages_json':messages_json,
            'settings':settings,
            'po':po
            }
        return render(request, self.template_name, context)
    
class ApprovePurchaseOrder(LoginRequiredMixin, PermissionRequiredMixin, View):
    template_name = 'purchase/approve.html'
    permission_required = 'inventory.view_purchase_order'

    def get(self, request, *args, **kwargs):
        now = timezone.datetime.now()
        messages_data = [{'messages':message.message,'tags':message.tags} for message in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)
        settings = Settings.objects.first()
        po = PurchaseOrder.objects.filter(status='approved')
        context = {
            'time':now,
            'year':now.year,
            'month':now.strftime('%B'),
            'day':now.strftime('%A'),
            'messages_json':messages_json,
            'settings':settings,
            'po':po
            }
        return render(request, self.template_name, context)
    
class approvedPurchaseOrderItems(View, PermissionRequiredMixin, LoginRequiredMixin):
    template_name = 'purchase/approvedItems.html'
    permission_required = ''
    
    def get(self, request, *args, **kwargs):
        now = timezone.datetime.now()
        messages_data = [{'messages':message.message,'tags':message.tags} for message in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)
        settings = Settings.objects.first()
        approved_items = PurchaseItem.objects.filter(purchase_order__status='approved').select_related(
            'purchase_order', 'vendor', 'material'
        )
        total_cost = approved_items.aggregate(total=Sum('price'))['total'] or 0
        context = {
            'time':now,
            'year':now.year,
            'month':now.strftime('%B'),
            'day':now.strftime('%A'),
            'messages_json':messages_json,
            'settings':settings,
            'approved_items': approved_items,
            'total_cost':total_cost
        }
        return render(request, self.template_name, context)
    
class cancelledPurchaseOrderItems(View, PermissionRequiredMixin, LoginRequiredMixin):
    template_name = 'purchase/cancelledItems.html'
    permission_required = ''
    
    def get(self, request, *args, **kwargs):
        now = timezone.datetime.now()
        messages_data = [{'messages':message.message,'tags':message.tags} for message in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)
        settings = Settings.objects.first()
        cancelled_items = PurchaseItem.objects.filter(purchase_order__status='cancelled').select_related(
            'purchase_order', 'vendor', 'material'
        )
        total_cost = cancelled_items.aggregate(total=Sum('price'))['total'] or 0
        context = {
            'time':now,
            'year':now.year,
            'month':now.strftime('%B'),
            'day':now.strftime('%A'),
            'messages_json':messages_json,
            'settings':settings,
            'cancelled_items': cancelled_items,
            'total_cost':total_cost
        }
        return render(request, self.template_name, context)


class PurchaseReport(LoginRequiredMixin, PermissionRequiredMixin, View):
    template_name = 'purchase/report.html'
    permission_required = 'purchase.view_purchasereport'

    def post(self, request, *args, **kwargs):
        start_date_str = request.POST.get('start_date', '').strip()
        end_date_str = request.POST.get('end_date', '').strip()
        export_format = request.POST.get('format', 'pdf').lower()

        # Validate input
        if not start_date_str or not end_date_str:
            return HttpResponse("No valid date was selected. Make sure you select both start and end dates.")

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse("Invalid date format. Use YYYY-MM-DD.")

        # Convert to datetime for filtering
        start_datetime = datetime.combine(start_date, time.min)
        end_datetime = datetime.combine(end_date, time.max)

        # Fetch approved and received orders within range
        purchase_orders = PurchaseOrder.objects.filter(
            status='approved',
            received_date__gte=start_datetime,
            received_date__lte=end_datetime
        ).prefetch_related('items__material')

        if not purchase_orders.exists():
            return HttpResponse("No orders found for the selected date range.")

        # Total purchase cost
        total_cost = sum(po.total_cost for po in purchase_orders)

        if export_format == 'excel':
            return self.generate_excel(purchase_orders, start_date_str, end_date_str, total_cost)

        # PDF export
        context = {
            'start_date': start_date_str,
            'end_date': end_date_str,
            'purchase_orders': purchase_orders,
            'total_cost': total_cost
        }
        html_template = render_to_string(self.template_name, context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="purchase_report.pdf"'

        result = pisa.pisaDocument(BytesIO(html_template.encode("UTF-8")), dest=response, link_callback=self.link_callback)
        if result.err:
            return HttpResponse(f"We had some error generating PDF: <pre>{str(result.err)}</pre>")

        return response

    def generate_excel(self, purchase_orders, start_date, end_date, total_cost):
        wb = Workbook()

        # --- Purchase Orders Sheet ---
        ws_orders = wb.active
        ws_orders.title = "Purchase Orders"

        ws_orders.append(['Purchase Orders Report'])
        ws_orders.append([f'Date Range: {start_date} to {end_date}'])
        ws_orders.append([])
        ws_orders.append(['Order #', 'Received Date', 'Total Cost'])

        for po in purchase_orders:
            ws_orders.append([
                po.order_number,
                po.received_date.strftime('%Y-%m-%d %H:%M') if po.received_date else 'Pending',
                float(po.total_cost)
            ])

        ws_orders.append([])
        ws_orders.append(['', '', 'Total Purchase Cost:', float(total_cost)])

        # Auto-fit columns
        for col in ws_orders.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
            ws_orders.column_dimensions[col[0].column_letter].width = max_length

        # --- Product Summary Sheet ---
        ws_summary = wb.create_sheet(title="Product Summary")
        ws_summary.append(['Product Summary'])
        ws_summary.append([f'Date Range: {start_date} to {end_date}'])
        ws_summary.append([])
        ws_summary.append(['Product', 'Quantity Purchased', 'Purchase Cost', 'COGS', 'Revenue', 'Profit', 'Profit Margin (%)'])

        # Aggregate product-level data
        product_data = {}
        for po in purchase_orders:
            for item in po.items.all():
                name = str(item.material.name)
                if name not in product_data:
                    product_data[name] = {'quantity': 0, 'cogs': 0, 'revenue': 0}
                product_data[name]['quantity'] += item.quantity
                product_data[name]['cogs'] += float(item.price)  # Cost of goods purchased
                product_data[name]['revenue'] += float(item.price)  # Placeholder; replace if you have actual revenue

        for name, metrics in product_data.items():
            profit = metrics['revenue'] - metrics['cogs']
            profit_margin = (profit / metrics['revenue'] * 100) if metrics['revenue'] > 0 else 0
            ws_summary.append([
                name,
                metrics['quantity'],
                metrics['cogs'],
                metrics['cogs'],  # COGS
                metrics['revenue'],
                profit,
                round(profit_margin, 2)
            ])

        # Auto-fit summary sheet
        for col in ws_summary.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
            ws_summary.column_dimensions[col[0].column_letter].width = max_length

        # Save Excel
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=purchase_report.xlsx'
        return response

    def link_callback(self, uri, rel):
        """Convert HTML URIs to absolute paths for xhtml2pdf."""
        if uri.startswith(settings.MEDIA_URL):
            path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
        elif uri.startswith(settings.STATIC_URL):
            path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
        else:
            return uri
        if not os.path.isfile(path):
            raise Exception(f'Media path does not exist: {path}')
        return path




# PRODUCTMANAGEMENT
class ProductManagement(LoginRequiredMixin, PermissionRequiredMixin, View):
    template_name = 'Product/management.html'
    permission_required = 'product.add_product'  # Adjust per your app

    def get(self, request, *args, **kwargs):
        settings = Settings.objects.first()
        products = Product.objects.select_related('product').all()
        materials = Material.objects.all()
        now = datetime.now()

        # Get Django messages as JSON for front-end
        messages_data = [{'message': message.message, 'tags': message.tags} for message in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)

        context = {
            'products': products,
            'time': now,
            'day': now.strftime('%A'),
            'month': now.strftime('%B'),
            'year': now.year,
            'messages_json': messages_json,
            'materials': materials,
            'settings': settings
        }

        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        product_ids = request.POST.getlist('product[]')
        quantities = request.POST.getlist('quantity[]')
        descriptions = request.POST.getlist('description[]')
        wholesale_prices = request.POST.getlist('wholesale_price[]')
        retail_prices = request.POST.getlist('retail_price[]')
        threshold_quantities = request.POST.getlist('threshold_quantity[]')
        user = request.user

        # Validate that all lists have the same length
        list_lengths = [len(product_ids), len(quantities), len(descriptions), len(wholesale_prices), len(retail_prices), len(threshold_quantities)]
        if len(set(list_lengths)) != 1:
            messages.error(request, "Mismatch in submitted form data lengths.")
            return redirect('product-management')

        try:
            with transaction.atomic():
                created_count = 0
                for i in range(len(product_ids)):
                    material = get_object_or_404(Material, id=product_ids[i])

                    # Skip if Product already exists
                    if Product.objects.filter(product=material).exists():
                        messages.warning(request, f"Product for material '{material.name}' already exists and was skipped.")
                        continue

                    # Convert numeric inputs
                    try:
                        quantity = int(quantities[i])
                        threshold_quantity = int(threshold_quantities[i])
                        wholesale_price = float(wholesale_prices[i])
                        retail_price = float(retail_prices[i])
                    except (ValueError, TypeError):
                        messages.error(request, f"Invalid numeric data for material '{material.name}'. Skipping.")
                        continue

                    # Create Product instance
                    product = Product(
                        product=material,
                        quantity=quantity,
                        description=descriptions[i],
                        wholesale_price=wholesale_price,
                        retail_price=retail_price,
                        threshold_quantity=threshold_quantity,
                        user=user
                    )

                    # Save product safely; will deduct material stock
                    try:
                        product.save()
                        created_count += 1
                    except ValueError as ve:
                        messages.error(request, str(ve))
                        continue

                if created_count > 0:
                    messages.success(request, f"{created_count} product(s) created successfully.")
                else:
                    messages.info(request, "No products were created.")

        except Exception as e:
            messages.error(request, f"An unexpected error occurred: {str(e)}")

        return redirect('product-management')




class UpdateDeleteProduct(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = ''

    def post(self, request, id, *args, **kwargs):
        product = get_object_or_404(Product, id=id)

        # # Store the old quantity for comparison
        # old_quantity = product.quantity

        # Get new values
        new_retail_price = request.POST['retail_price']
        new_wholesale_price = request.POST['wholesale_price']
        new_quantity = int(request.POST['quantity'])

        # Update product
        product.retail_price = new_retail_price
        product.wholesale_price = new_wholesale_price
        product.quantity = new_quantity
        product.save()  # This will now trigger your stock adjustment

        messages.success(request, 'Product updated successfully')
        return redirect('product-management')

    def get(self, request, id, *args, **kwargs):
        Product.objects.get(id=id).delete()
        messages.success(request, 'Product deleted successfully')
        return redirect('product-management')




# OrderManagement
class OrderManagementView(LoginRequiredMixin, PermissionRequiredMixin, View):
    template_name = 'order/management.html'
    permission_required = 'inventory.view_order'

    def get(self, request, *args, **kwargs):
        now = timezone.now()
        settings = Settings.objects.first()
        orders = Order.objects.all().select_related('user')
        products = Product.objects.select_related('product').all()

        messages_data = [{'message': m.message, 'tags': m.tags} for m in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)

        context = {
            'time': now,
            'year': now.year,
            'month': now.strftime('%B'),
            'day': now.strftime('%A'),
            'messages_json': messages_json,
            'settings': settings,
            'products': products,
            'orders': orders
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        order_type = request.POST.get('order_type')
        product_ids = request.POST.getlist('product_id[]')
        quantities = request.POST.getlist('quantity[]')
        user = request.user

        # Validate input
        if not order_type or order_type not in ['retail', 'wholesale']:
            messages.error(request, "Invalid order type.")
            return redirect('order-management')

        if not product_ids or not quantities or len(product_ids) != len(quantities):
            messages.error(request, "Missing or mismatched order item data.")
            return redirect('order-management')

        try:
            with transaction.atomic():
                # ✅ Create the order (order_id auto-generated in model's save())
                order = Order.objects.create(
                    order_type=order_type,
                    status='pending',
                    user=user
                )

                created_items = 0
                for i in range(len(product_ids)):
                    product = get_object_or_404(Product, id=product_ids[i])

                    # Convert quantity and validate
                    try:
                        quantity = int(quantities[i])
                        if quantity <= 0:
                            raise ValueError(f"Quantity must be positive for product {product.product.name}.")
                        if quantity > product.quantity:
                            raise ValueError(f"Not enough stock for product {product.product.name}. Available: {product.quantity}, requested: {quantity}")
                    except ValueError as ve:
                        messages.error(request, str(ve))
                        continue  # skip this item but continue with others

                    # Create OrderItem (this will trigger update_total in its save())
                    OrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity=quantity,
                        user=user
                    )
                    created_items += 1

                if created_items == 0:
                    # No valid items created, delete the empty order
                    order.delete()
                    messages.error(request, "No valid order items were created. Order cancelled.")
                else:
                    # Optional: recalc total for safety
                    order.update_total()
                    messages.success(request, f"Order {order.order_id} created with {created_items} item(s).")

        except Exception as e:
            messages.error(request, f"Failed to create order: {str(e)}")

        return redirect('order-management')




class ApproveOrderManagementView(View, PermissionRequiredMixin, LoginRequiredMixin):
    permission_required = ''

    def get(self, request, id, *args, **kwargs):
        order = get_object_or_404(Order, id=id)
        if order.status != 'approved':
            order.status = 'approved'
            order.approved_date = timezone.now()
            order.save()
            messages.success(request, 'order with ID f{oredr.order_id} approved successfully')
            return redirect('order-management')




class CancellOrderManagementView(View, PermissionRequiredMixin, LoginRequiredMixin):
    permission_required = ''

    def get(self, request, id, *args, **kwargs):
        order = get_object_or_404(Order, id=id)
        if order.status != 'cancelled':
            order.status = 'cancelled'
            order.cancelled_date = datetime.now()
            order.save()
            messages.success(request, 'order with ID f{oredr.order_id} cancelled successfully')
            return redirect('order-management')




class OrderViewDetail(View, LoginRequiredMixin, PermissionRequiredMixin):
    template_name = 'order/detail.html'
    permission_required = 'inventory.view_order'

    def get(self, request, id, *args, **kwargs):
        order = Order.objects.get(id=id)
        now = timezone.datetime.now()
        messages_data = [{'messages':message.message,'tags':message.tags} for message in messages.get_messages(request)]
        messages_json = json.dumps(messages_data)
        settings = Settings.objects.first()
        context = {
            'time':now,
            'year':now.year,
            'month':now.strftime('%B'),
            'day':now.strftime('%A'),
            'messages_json':messages_json,
            'settings':settings,
            'order':order,
        }
        return render(request, self.template_name, context)





class UpdateDeleteOrder(View, LoginRequiredMixin, PermissionRequiredMixin):
    permission_required = 'inventory.change_order'

    def get(self, request,*args, **kwargs):
        order_id = kwargs.get('pk')
        try:
            order = Order.objects.get(id=order_id).delete()
            messages.success(request, 'order deleted successfully')
            return redirect('order-management')
        except Order.DoesNotExist:
            messages.error(request, "Order not found.")
            return redirect('order-management')

    def post(self, request, *args, **kwargs):
            order_id = request.POST.get('order_id')
            order = get_object_or_404(Order, id=order_id, user=request.user)

            # Update order fields
            order_type = request.POST.get('order_type')
            status = request.POST.get('status')  # If editable

            if order_type:
                order.order_type = order_type
            if status:
                order.status = status

            order.save()  # Save basic updates

            # Retrieve submitted data
            item_ids = request.POST.getlist('item_id[]')  # optional; may not exist for new rows
            product_ids = request.POST.getlist('product_id[]')
            quantities = request.POST.getlist('quantity[]')

            # Track existing item IDs to retain
            retained_item_ids = []

            for i in range(len(product_ids)):
                try:
                    quantity = int(quantities[i])
                    if quantity <= 0:
                        continue  # skip invalid quantities

                    product = Product.objects.get(id=product_ids[i])

                    if item_ids and i < len(item_ids) and item_ids[i]:
                        # Update existing item
                        item = OrderItem.objects.get(id=item_ids[i], order=order)
                        item.product = product
                        item.quantity = quantity
                        item.save()
                        retained_item_ids.append(item.id)
                    else:
                        # New item
                        item = OrderItem.objects.create(
                            order=order,
                            product=product,
                            quantity=quantity
                        )
                        retained_item_ids.append(item.id)
                except Exception as e:
                    # Handle error or log
                    print(f"Error processing item: {e}")
                    continue

            # Remove deleted items
            OrderItem.objects.filter(order=order).exclude(id__in=retained_item_ids).delete()

            # Recalculate total
            order.update_total()      
            return redirect('order-management')  




class SingleOrderReport(View, LoginRequiredMixin, PermissionRequiredMixin):
    template_name = 'order/detail_report.html'
    permission_required = ''

    def get(self, request, id, *args, **kwargs):
            # 1. Fetch the PurchaseOrder or return 404
            order = get_object_or_404(Order, id=id)

            # 2. Render HTML template with context
            context = {
                'order': order,
                'request': request,  # required for absolute static/media URLs
            }
            html = render_to_string(self.template_name, context)

            # 3. Create PDF from HTML
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="Order-Invoice.pdf"'

            # Create a BytesIO buffer
            result = BytesIO()

            # Convert HTML to PDF
            pdf = pisa.pisaDocument(BytesIO(html.encode('utf-8')), dest=result,
                encoding='UTF-8', link_callback=self.link_callback)

            if pdf.err:
                return HttpResponse(f"Error generating PDF: <pre>{pdf.err}</pre>")

            # 4. Write result to response
            response.write(result.getvalue())
            return response

    def link_callback(self, uri, rel):
            """
            Convert HTML URIs to absolute system paths so xhtml2pdf can access those resources (e.g. images, CSS).
            """
            if uri.startswith(settings.MEDIA_URL):
                path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
            elif uri.startswith(settings.STATIC_URL):
                path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
            else:
                return uri  # return unchanged

            if not os.path.isfile(path):
                raise Exception(f'Media path does not exist: {path}')
            return path





class OrderReport(LoginRequiredMixin, PermissionRequiredMixin, View):
    template_name = 'Order/report.html'
    permission_required = ''  # Set your permission

    def post(self, request, *args, **kwargs):
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        export_format = request.POST.get('format', 'pdf')  # 'pdf' by default

        if not start_date_str or not end_date_str:
            return HttpResponse('No valid date was selected.')

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

            start_datetime = datetime.combine(start_date, datetime.min.time())
            end_datetime = datetime.combine(end_date + timedelta(days=1), datetime.min.time())

            orders = Order.objects.filter(
                status='approved',
                approved_date__gte=start_datetime,
                approved_date__lt=end_datetime
            ).prefetch_related('items__product')  # Optimization

            total_price = orders.aggregate(Sum('total_price'))['total_price__sum'] or 0

            if export_format.lower() == 'excel':
                return self.generate_excel(orders, start_date_str, end_date_str, total_price)

            # PDF export
            context = {
                'start_date': start_date_str,
                'end_date': end_date_str,
                'orders': orders,
                'total_price': total_price
            }

            html_template = render_to_string(self.template_name, context)
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="order_report.pdf"'

            result = pisa.pisaDocument(BytesIO(html_template.encode("UTF-8")), dest=response, link_callback=self.link_callback)
            if result.err:
                return HttpResponse(f"We had some error generating PDF: <pre>{str(result.err)}</pre>")

            return response

        except ValueError:
            return HttpResponse('Invalid date format.')

    def generate_excel(self, purchase_orders, start_date, end_date, total_cost):
        wb = Workbook()
        ws = wb.active
        ws.title = "Order Report"

        # Report title and date range
        ws.append(['Order Report'])
        ws.append([f'Date Range: {start_date} to {end_date}'])
        ws.append([f"Statuses: {', '.join([po.status for po in purchase_orders])}"])
        ws.append([])

        # Column headers
        ws.append(['Order ID'])

        for po in purchase_orders:
            ws.append([po.order_id])
            ws.append(['', 'PRODUCT', 'QUANTITY(qts)', 'UNIT_PRICE(p)', 'TOTAL'])

            for item in po.items.all():
                total_item_price = item.quantity * item.unit_price
                ws.append([
                    '',
                    str(item.product.product),
                    item.quantity,
                    float(item.unit_price),
                    float(total_item_price)
                ])

            ws.append([])

        ws.append([])
        ws.append(['', '', '', 'Total Order Price:', float(total_cost)])

        # Auto-fit columns
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells if cell.value) + 2
            ws.column_dimensions[column_cells[0].column_letter].width = max_length

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=order_report.xlsx'
        return response

    def link_callback(self, uri, rel):
        """
        Convert HTML URIs to absolute system paths for xhtml2pdf.
        """
        if uri.startswith(settings.MEDIA_URL):
            path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
        elif uri.startswith(settings.STATIC_URL):
            path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
        else:
            return uri

        if not os.path.isfile(path):
            raise Exception(f'Media path does not exist: {path}')
        return path





class FinancialReportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'order.generate_report'

    def post(self, request, *args, **kwargs):

        filter_type = request.POST.get('filter_type', 'date')
        export_format = request.POST.get('format', 'excel').lower()

        # ================= DATE FILTER LOGIC =================
        if filter_type == "year":
            year = request.POST.get('year')

            if not year:
                return HttpResponse("Please select a year.")

            try:
                year = int(year)
                start_date = datetime(year, 1, 1).date()
                end_date = datetime(year, 12, 31).date()
            except ValueError:
                return HttpResponse("Invalid year selected.")

        else:
            start_date_str = request.POST.get('start_date')
            end_date_str = request.POST.get('end_date')

            if not start_date_str or not end_date_str:
                return HttpResponse("Please provide valid start and end dates.")

            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError:
                return HttpResponse("Invalid date format. Use YYYY-MM-DD.")

        # ================= EXPENSES =================
        expenses_queryset = Expenses.objects.filter(
            date__range=(start_date, end_date)
        )

        expenses_data = [
            {
                'name': s['expenses_name__name'],
                'total': s['total'] or Decimal('0.00')
            }
            for s in expenses_queryset
            .values('expenses_name__name')
            .annotate(total=Sum('cost'))
        ]

        grand_expense_total = sum(
            [e['total'] for e in expenses_data],
            Decimal('0.00')
        )

        # ================= REVENUE =================
        wholesale_data = self.get_revenue_data(start_date, end_date, 'wholesale')
        retail_data = self.get_revenue_data(start_date, end_date, 'retail')

        combined_data = wholesale_data + retail_data

        total_revenue = sum(
            [p['revenue'] for p in combined_data],
            Decimal('0.00')
        )

        total_cogs = sum(
            [p['cogs'] for p in combined_data],
            Decimal('0.00')
        )

        total_profit = total_revenue - total_cogs
        net_profit = total_profit - grand_expense_total

        # ================= PDF EXPORT =================
        if export_format == 'pdf':

            context = {
                "start_date": start_date,
                "end_date": end_date,
                "expenses": expenses_data,
                "grand_total": grand_expense_total,
                "wholesale": wholesale_data,
                "retail": retail_data,
                "total_revenue": total_revenue,
                "total_cogs": total_cogs,
                "total_profit": total_profit,
                "total_expenses": grand_expense_total,
                "net_profit": net_profit
            }

            html = render_to_string("final/report.html", context)

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="financial_report.pdf"'

            pisa_status = pisa.CreatePDF(
                html,
                dest=response,
                link_callback=self.link_callback
            )

            if pisa_status.err:
                return HttpResponse(f"Error generating PDF: {pisa_status.err}")

            return response

        # ================= EXCEL EXPORT =================
        wb = Workbook()

        self.create_expenses_sheet(wb, expenses_data, grand_expense_total)
        self.create_revenue_sheet(wb, wholesale_data, "Wholesale Revenue")
        self.create_revenue_sheet(wb, retail_data, "Retail Revenue")
        self.create_summary_sheet(
            wb,
            total_revenue,
            total_cogs,
            total_profit,
            grand_expense_total,
            net_profit
        )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="financial_report.xlsx"'

        return response

    # ================= AVERAGE COST =================
    def get_average_unit_cost(self, product):

        purchase_items = PurchaseItem.objects.filter(material=product)
        total_qty = purchase_items.aggregate(total=Sum('quantity'))['total'] or 0

        if total_qty == 0:
            return Decimal('0.00')

        total_cost = sum([pi.unit_price * pi.quantity for pi in purchase_items])
        return total_cost / total_qty

    # ================= REVENUE DATA =================
    def get_revenue_data(self, start_date, end_date, order_type):

        orders = Order.objects.filter(
            status='approved',
            order_type=order_type,
            approved_date__date__range=(start_date, end_date)
        ).prefetch_related('items__product__product')

        product_data = {}

        for order in orders:
            for item in order.items.all():

                product_name = item.product.product.name
                qty = item.quantity
                purchase_price = self.get_average_unit_cost(item.product.product)

                selling_price = (
                    item.product.wholesale_price
                    if order_type == 'wholesale'
                    else item.product.retail_price
                )

                cogs = purchase_price * qty
                revenue = selling_price * qty
                profit = revenue - cogs

                if product_name not in product_data:
                    product_data[product_name] = {
                        'name': product_name,
                        'quantity': 0,
                        'purchase_price': purchase_price,
                        'selling_price': selling_price,
                        'cogs': Decimal('0.00'),
                        'revenue': Decimal('0.00'),
                        'profit': Decimal('0.00')
                    }

                product_data[product_name]['quantity'] += qty
                product_data[product_name]['cogs'] += cogs
                product_data[product_name]['revenue'] += revenue
                product_data[product_name]['profit'] += profit

        return list(product_data.values())

    # ================= EXCEL SHEETS =================
    def create_expenses_sheet(self, wb, expenses_data, grand_total):

        ws = wb.active
        ws.title = "Expenses"

        ws.append(["Service", "Total Cost"])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for e in expenses_data:
            ws.append([e['name'], float(e['total'])])

        ws.append(["Grand Total", float(grand_total)])
        ws[f"A{ws.max_row}"].font = Font(bold=True)
        ws[f"B{ws.max_row}"].font = Font(bold=True)

        if ws.max_row > 1:
            tab = Table(displayName="ExpensesTable", ref=f"A1:B{ws.max_row}")
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(tab)

    def create_revenue_sheet(self, wb, data, title):

        ws = wb.create_sheet(title)

        ws.append([
            "Product Name",
            "Quantity Sold",
            "Purchase Price",
            "Selling Price",
            "COGS",
            "Revenue",
            "Profit"
        ])

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for p in data:
            ws.append([
                p['name'],
                p['quantity'],
                float(p['purchase_price']),
                float(p['selling_price']),
                float(p['cogs']),
                float(p['revenue']),
                float(p['profit'])
            ])

        if ws.max_row > 1:
            tab = Table(
                displayName=title.replace(" ", "") + "Table",
                ref=f"A1:G{ws.max_row}"
            )
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(tab)

    def create_summary_sheet(self, wb, total_revenue, total_cogs, total_profit, total_expenses, net_profit):

        ws = wb.create_sheet("Financial Summary")

        ws.append(["Metric", "Amount"])
        for cell in ws[1]:
            cell.font = Font(bold=True)

        rows = [
            ("Total Revenue", total_revenue),
            ("Total COGS", total_cogs),
            ("Total Profit", total_profit),
            ("Total Expenses", total_expenses),
            ("Net Profit", net_profit)
        ]

        for r in rows:
            ws.append([r[0], float(r[1])])

        if ws.max_row > 1:
            tab = Table(displayName="FinancialSummaryTable", ref=f"A1:B{ws.max_row}")
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
            ws.add_table(tab)

    # ================= STATIC HANDLER =================
    def link_callback(self, uri, rel):

        if uri.startswith(settings.MEDIA_URL):
            path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))

        elif uri.startswith(settings.STATIC_URL):
            path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))

        else:
            return uri

        if not os.path.isfile(path):
            raise Exception(f'Media path does not exist: {path}')

        return path





class UnsoldStockReportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'order.generate_report'

    def get(self, request, *args, **kwargs):
        # Generate report without any date/year filter
        product_data = {}

        # All purchased quantities
        purchases = PurchaseItem.objects.all()
        for item in purchases:
            product = item.material
            name = product.name
            qty = item.quantity
            cost = item.unit_price * qty

            if name not in product_data:
                product_data[name] = {
                    "name": name,
                    "purchased_qty": 0,
                    "sold_qty": 0,
                    "total_purchase_cost": Decimal("0.00")
                }

            product_data[name]["purchased_qty"] += qty
            product_data[name]["total_purchase_cost"] += cost

        # All sold quantities
        orders = Order.objects.filter(status="approved").prefetch_related("items__product__product")
        for order in orders:
            for item in order.items.all():
                product = item.product.product
                name = product.name
                qty = item.quantity

                if name not in product_data:
                    product_data[name] = {
                        "name": name,
                        "purchased_qty": 0,
                        "sold_qty": 0,
                        "total_purchase_cost": Decimal("0.00")
                    }

                product_data[name]["sold_qty"] += qty

        # Calculate unsold and COGS
        final_data = []
        total_unsold_value = Decimal("0.00")

        for product in product_data.values():
            purchased = product["purchased_qty"]
            sold = product["sold_qty"]
            unsold = purchased - sold

            if purchased > 0:
                avg_cost = product["total_purchase_cost"] / purchased
            else:
                avg_cost = Decimal("0.00")

            unsold_cogs = unsold * avg_cost
            total_unsold_value += unsold_cogs

            final_data.append({
                "name": product["name"],
                "purchased_qty": purchased,
                "sold_qty": sold,
                "unsold_qty": unsold,
                "avg_cost": avg_cost,
                "unsold_cogs": unsold_cogs
            })

        # Render PDF
        context = {
            "products": final_data,
            "total_unsold_value": total_unsold_value
        }

        html = render_to_string("final/stock.html", context)
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = "attachment; filename=unsold_stock_report.pdf"

        pisa_status = pisa.CreatePDF(html, dest=response, link_callback=self.link_callback)
        if pisa_status.err:
            return HttpResponse("Error generating PDF")

        return response

    # Handle static files in PDF
    def link_callback(self, uri, rel):
        if uri.startswith(settings.MEDIA_URL):
            path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
        elif uri.startswith(settings.STATIC_URL):
            path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
        else:
            return uri

        if not os.path.isfile(path):
            raise Exception(f'Media path does not exist: {path}')
        return path